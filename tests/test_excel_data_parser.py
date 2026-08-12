"""Tests for ExcelDataParser: workbook creation, templating, and data writes.

Uses pytest's `tmp_path` fixture instead of a hardcoded path so these tests
run identically on any machine/CI runner and never touch the real repo.
"""

from openpyxl import load_workbook

from reviewbygpt.lib.excel_data_parser import ExcelDataParser


def test_create_excel_file(tmp_path):
    parser = ExcelDataParser(str(tmp_path))
    assert parser.create_excel_file() is True
    assert (tmp_path / "analysed.xlsx").exists()


def test_apply_excel_template_writes_headers(tmp_path):
    parser = ExcelDataParser(str(tmp_path))
    parser.create_excel_file()

    parser.apply_excel_template("qa_sheet", ["TITLE", "QE1", "QE1_SCORE", "TOTAL_SCORE"])

    wb = load_workbook(tmp_path / "analysed.xlsx")
    ws = wb["qa_sheet"]
    headers = [cell.value for cell in ws[1]]
    assert headers == ["TITLE", "QE1", "QE1_SCORE", "TOTAL_SCORE"]


def test_fill_excel_with_data_appends_row(tmp_path):
    parser = ExcelDataParser(str(tmp_path))
    parser.create_excel_file()
    parser.apply_excel_template("qa_sheet", ["TITLE", "QE1_SCORE", "TOTAL_SCORE"])

    ok = parser.fill_excel_with_data(
        "qa_sheet", {"TITLE": "My Paper", "QE1_SCORE": 1.0, "TOTAL_SCORE": 6.5}
    )
    assert ok is True

    rows = parser.get_existing_data("qa_sheet")
    assert len(rows) == 1
    assert rows[0]["TITLE"] == "My Paper"
    assert rows[0]["TOTAL_SCORE"] == 6.5


def test_get_existing_data_on_missing_sheet_returns_empty_list(tmp_path):
    parser = ExcelDataParser(str(tmp_path))
    parser.create_excel_file()
    assert parser.get_existing_data("does_not_exist") == []
