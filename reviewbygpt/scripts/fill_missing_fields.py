import os
import logging
import pandas as pd
import re
from openpyxl import load_workbook
from reviewbygpt.lib.review_data_parser import ReviewDataParser
from reviewbygpt.lib.excel_data_parser import ExcelDataParser

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class MissingFieldsFiller:
    def __init__(self, excel_directory, response_directory, review_config):
        """
        Initialize the missing fields filler
        
        Args:
            excel_directory (str): Directory containing the analysed.xlsx file
            response_directory (str): Directory containing the response txt files
            review_config (str): Path to the review configuration file
        """
        # Define special field names that might cause issues
        self.special_field_names = [
            "REAL-TIME FEEDBACK & CONTROL",
            "REAL-TIME FEEDBACK",
            "EFFICIENCY & ACCURACY",
            "REAL-TIME & MOTION PLANNING",
            "HARDWARE & SOFTWARE",
            "CHALLENGES & LIMITATIONS"
        ]
        self.excel_directory = excel_directory
        self.response_directory = response_directory
        self.review_config = review_config
        
        # Initialize parsers
        self.review_parser = ReviewDataParser(review_config)
        self.excel_parser = ExcelDataParser(excel_directory, review_config)
        
        # Excel file path
        self.excel_file = os.path.join(excel_directory, "analysed.xlsx")
        
        logger.info(f"Missing Fields Filler initialized with Excel directory: {excel_directory}")
        logger.info(f"Response files directory: {response_directory}")
    
    def read_excel_data(self):
        """
        Read data from the Excel file to identify papers and missing fields
        
        Returns:
            tuple: (de_data, qa_data) - Data extraction and quality assessment data
        """
        try:
            # Check if file exists
            if not os.path.exists(self.excel_file):
                logger.error(f"Excel file not found: {self.excel_file}")
                return None, None
            
            # Get existing data from sheets
            de_data = self.excel_parser.get_existing_data("de_sheet")
            qa_data = self.excel_parser.get_existing_data("qa_sheet")
            
            logger.info(f"Successfully read {len(de_data)} papers from de_sheet")
            logger.info(f"Successfully read {len(qa_data)} papers from qa_sheet")
            
            return de_data, qa_data
        except Exception as e:
            logger.error(f"Error reading Excel data: {e}")
            return None, None
    
    def find_response_file(self, paper_title):
        """
        Find the response file for a given paper title
        
        Args:
            paper_title (str): Title of the paper
            
        Returns:
            str: Path to response file or None if not found
        """
        try:
            # Extract some identifying words from the title to find a filename match
            # This uses the first few significant words to make a more reliable match
            title_words = re.findall(r'\b\w+\b', paper_title.lower())
            significant_words = [word for word in title_words if len(word) > 3][:3]  # First 3 significant words
            
            # List all response files
            for filename in os.listdir(self.response_directory):
                if filename.endswith("_response.txt"):
                    # Check if the filename contains the paper ID
                    file_path = os.path.join(self.response_directory, filename)
                    
                    # Read the response file to check if it contains the paper title
                    with open(file_path, 'r', encoding='utf-8') as f:
                        content = f.read()
                        
                        # Check if title appears in the response
                        if all(word in content.lower() for word in significant_words):
                            logger.info(f"Found response file for '{paper_title}': {filename}")
                            return file_path
            
            logger.warning(f"No response file found for paper: {paper_title}")
            return None
        except Exception as e:
            logger.error(f"Error finding response file for '{paper_title}': {e}")
            return None
    
    def extract_missing_fields(self, paper_data, response_file):
        """
        Extract missing fields from a response file
        
        Args:
            paper_data (dict): Current paper data with possible missing fields
            response_file (str): Path to response file
            
        Returns:
            dict: Updated paper data with missing fields filled in
        """
        try:
            # Read the response file
            with open(response_file, 'r', encoding='utf-8') as f:
                response_content = f.read()
            
            # Extract all data fields from the response
            all_de_data = self.review_parser.get_data_extraction_text(response_content)
            all_qa_data = self.review_parser.get_quality_assessment_text(response_content)
            
            # Look for special field names manually if they're missing
            for special_field in self.special_field_names:
                if special_field not in all_de_data:
                    # Try to find the field manually in the response content
                    pattern = re.compile(rf"{re.escape(special_field)}\s*:\s*(.*?)(?=\s+[A-Z][A-Z\s&-]+:|==DATA_EXTRACTION_END==|$)", re.DOTALL)
                    match = pattern.search(response_content)
                    if match:
                        value = match.group(1).strip()
                        all_de_data[special_field] = value
                        logger.info(f"Manually extracted special field '{special_field}': {value}")
            
            # Create a copy of the paper data
            updated_data = paper_data.copy()
            
            # Find and fill missing fields
            for field, value in all_de_data.items():
                if field not in updated_data or not updated_data[field]:
                    if value:  # Only update if we have a non-empty value
                        updated_data[field] = value
                        logger.info(f"Added missing field '{field}': {value}")
            
            return updated_data
        except Exception as e:
            logger.error(f"Error extracting missing fields: {e}")
            return paper_data
    
    def update_excel_file(self, sheet_name, paper_idx, updated_data):
        """
        Update a specific paper's data in the Excel file
        
        Args:
            sheet_name (str): Sheet name to update
            paper_idx (int): Index of the paper in the sheet
            updated_data (dict): Updated paper data
            
        Returns:
            bool: True if update was successful, False otherwise
        """
        try:
            # Load the workbook and select the sheet
            wb = load_workbook(self.excel_file)
            
            if sheet_name not in wb.sheetnames:
                logger.error(f"Sheet '{sheet_name}' not found in Excel file")
                return False
            
            ws = wb[sheet_name]
            
            # Get the headers from the first row
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(cell.value)
            
            # Update the specific row
            row_num = paper_idx + 2  # Adding 2 because Excel is 1-indexed and row 1 is headers
            
            # Check if row exists
            if row_num > ws.max_row:
                logger.error(f"Row {row_num} does not exist in sheet '{sheet_name}'")
                return False
            
            # Update each field in the row
            updates_made = False
            for col_idx, header in enumerate(headers, start=1):
                if header in updated_data and (ws.cell(row=row_num, column=col_idx).value != updated_data[header]):
                    ws.cell(row=row_num, column=col_idx, value=updated_data[header])
                    updates_made = True
            
            # Save the workbook if updates were made
            if updates_made:
                wb.save(self.excel_file)
                logger.info(f"Updated paper in row {row_num} of sheet '{sheet_name}'")
                return True
            else:
                logger.info(f"No updates needed for paper in row {row_num}")
                return True
            
        except Exception as e:
            logger.error(f"Error updating Excel file: {e}")
            return False
    
    def process_all_papers(self):
        """
        Process all papers in the Excel file to fill missing fields
        
        Returns:
            tuple: (success_count, total_count) - Number of successfully processed papers and total papers
        """
        # Read data from Excel
        de_data, qa_data = self.read_excel_data()
        
        if not de_data:
            logger.error("No data extraction data found in Excel")
            return 0, 0
        
        success_count = 0
        total_count = len(de_data)
        
        # Process each paper
        for idx, paper in enumerate(de_data):
            try:
                # Check if paper has a title
                if "TITLE" not in paper or not paper["TITLE"]:
                    logger.warning(f"Paper at index {idx} has no title, skipping")
                    continue
                
                paper_title = paper["TITLE"]
                logger.info(f"Processing paper: {paper_title}")
                
                # Check for missing fields
                missing_fields = []
                for key in paper:
                    if not paper[key]:
                        missing_fields.append(key)
                
                # Skip if no missing fields
                if not missing_fields:
                    logger.info(f"No missing fields for paper: {paper_title}")
                    success_count += 1
                    continue
                
                logger.info(f"Found {len(missing_fields)} missing fields: {', '.join(missing_fields)}")
                
                # Find response file
                response_file = self.find_response_file(paper_title)
                if not response_file:
                    logger.warning(f"No response file found for paper: {paper_title}, skipping")
                    continue
                
                # Extract missing fields
                updated_data = self.extract_missing_fields(paper, response_file)
                
                # Update Excel
                if self.update_excel_file("de_sheet", idx, updated_data):
                    success_count += 1
                
            except Exception as e:
                logger.error(f"Error processing paper at index {idx}: {e}")
        
        logger.info(f"Successfully processed {success_count} out of {total_count} papers")
        return success_count, total_count

def main():
    """
    Main function to run the missing fields filler
    """
    # Configuration
    excel_directory = "/home/pedrodias/Documents/phd/systematic-review-disassebly/sheet"  # Directory containing analysed.xlsx
    response_directory = "/home/pedrodias/Documents/phd/systematic-review-disassebly/debug_logs"  # Directory containing response txt files
    review_config = "/home/pedrodias/Documents/git-repos/ReviewbyGPT/config/review_data.yaml"  # Path to review configuration
    
    # Initialize and run the filler
    filler = MissingFieldsFiller(excel_directory, response_directory, review_config)
    success_count, total_count = filler.process_all_papers()
    
    logger.info(f"Processing complete. Successfully updated {success_count} out of {total_count} papers.")

if __name__ == "__main__":
    main()