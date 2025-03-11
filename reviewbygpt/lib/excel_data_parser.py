import os
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
import shutil

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ExcelDataParser:
    def __init__(self, excel_file_path):
        """
        Initialize the Excel data parser with proper path handling
        """
        # Store the base directory
        self.excel_dir = excel_file_path
        
        # Define the output Excel filename
        self.excel_filename = "analysed.xlsx"
        
        # Create the full path to the Excel file
        self.full_excel_path = os.path.join(self.excel_dir, self.excel_filename)
        
        logger.info(f"Excel parser initialized with file: {self.full_excel_path}")

    def create_excel_file(self):
        """
        Create a new Excel file with robust error handling
        """
        try:
            # Check if folder exists, create if needed
            if not os.path.exists(self.excel_dir):
                os.makedirs(self.excel_dir)
                logger.info(f"Created directory: {self.excel_dir}")

            # Create a fresh workbook
            wb = Workbook()
            
            # Keep the default sheet but rename it to avoid deletion errors
            if "Sheet" in wb.sheetnames:
                sheet = wb["Sheet"]
                sheet.title = "Info"
                # Add some metadata to this sheet
                sheet['A1'] = "Created on:"
                sheet['B1'] = f"{__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                sheet['A2'] = "File path:"
                sheet['B2'] = f"{self.full_excel_path}"
            
            # Save the workbook
            wb.save(self.full_excel_path)
            logger.info(f"Created new Excel file: {self.full_excel_path}")
            return True
        except Exception as e:
            logger.error(f"Error creating Excel file: {e}")
            return False

    def _backup_excel_file(self):
        """
        Create a backup of the Excel file if it exists
        """
        if os.path.exists(self.full_excel_path):
            backup_path = f"{self.full_excel_path}.bak"
            try:
                shutil.copy2(self.full_excel_path, backup_path)
                logger.info(f"Created backup of Excel file at: {backup_path}")
                return True
            except Exception as e:
                logger.warning(f"Failed to create backup: {e}")
        return False

    def _ensure_valid_workbook(self):
        """
        Ensure we have a valid workbook, creating a new one if necessary
        """
        try:
            # Try to load the workbook to test if it's valid
            if os.path.exists(self.full_excel_path):
                try:
                    wb = load_workbook(self.full_excel_path)
                    # If we get here, the workbook is valid
                    return wb
                except Exception as e:
                    logger.error(f"Excel file appears to be corrupted: {e}")
                    # Backup the corrupted file first
                    corrupted_path = f"{self.full_excel_path}.corrupted"
                    try:
                        shutil.move(self.full_excel_path, corrupted_path)
                        logger.info(f"Moved corrupted file to: {corrupted_path}")
                    except:
                        # If we can't move it, try to delete it
                        try:
                            os.remove(self.full_excel_path)
                            logger.info(f"Removed corrupted Excel file")
                        except Exception as del_err:
                            logger.error(f"Failed to clean up corrupted file: {del_err}")
            
            # Create a new workbook
            self.create_excel_file()
            return load_workbook(self.full_excel_path)
            
        except Exception as e:
            logger.error(f"Failed to ensure valid workbook: {e}")
            # Last resort - create a completely new file
            try:
                wb = Workbook()
                wb.save(self.full_excel_path)
                return wb
            except:
                raise RuntimeError("Cannot create or load Excel workbook")

    def apply_excel_template(self, sheet_name, identifiers):
        """
        Apply initial template to Excel sheet with better error handling
        """
        try:
            # Get a valid workbook
            wb = self._ensure_valid_workbook()
            
            # Check if the sheet exists, if not create it
            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
                logger.info(f"Created new sheet: {sheet_name}")
            else:
                ws = wb[sheet_name]
            
            # Check if this is a QA sheet by examining identifiers
            is_qa_sheet = any(identifier.startswith("QE") for identifier in identifiers)
            
            # For QA sheet, make sure 'Title' is included in the identifiers
            if is_qa_sheet and "TITLE" not in identifiers:
                identifiers = ["TITLE"] + identifiers
                logger.info("Added 'Title' field to QA sheet identifiers")
            
            # Apply the headers if the sheet is empty or force update
            if ws.max_row <= 1:  # Sheet is empty or has only headers
                # Set header style
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                
                for col_idx, identifier in enumerate(identifiers, start=1):
                    cell = ws.cell(row=1, column=col_idx, value=identifier)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Auto-adjust column widths based on identifier length
                for col_idx, identifier in enumerate(identifiers, start=1):
                    col_letter = ws.cell(row=1, column=col_idx).column_letter
                    width = max(15, len(str(identifier)) + 5)  # Minimum width of 15
                    ws.column_dimensions[col_letter].width = width
                
                logger.info(f"Applied template to sheet '{sheet_name}' with {len(identifiers)} identifiers")
            
            # Save the workbook
            wb.save(self.full_excel_path)
            return True
            
        except Exception as e:
            logger.error(f"Error applying template to sheet '{sheet_name}': {e}")
            # Attempt recovery by creating a completely new file
            try:
                logger.info("Attempting recovery by creating a new Excel file...")
                os.remove(self.full_excel_path)
                return self.apply_excel_template(sheet_name, identifiers)
            except:
                logger.error("Recovery failed")
                return False
    
    def fill_excel_with_data(self, sheet_name, data):
        """
        Fill Excel sheet with data with better error handling
        """
        try:
            # Try to load the existing workbook
            wb = self._ensure_valid_workbook()
            
            # Check if the sheet exists, create it if not
            if sheet_name not in wb.sheetnames:
                logger.warning(f"Sheet '{sheet_name}' not found, creating it")
                ws = wb.create_sheet(sheet_name)
                # We need to initialize the headers too
                if sheet_name == "qa_sheet":
                    # Create headers for QA sheet
                    # In apply_excel_template method, change:
                    qa_headers = ["Title"] + [f"QE{i}" for i in range(1, 9)] + [f"QE{i}_SCORE" for i in range(1, 9)] + ["TOTAL_SCORE"]
                    self.apply_excel_template(sheet_name, qa_headers)
                    # Reload the workbook
                    wb = load_workbook(self.full_excel_path)
                    ws = wb[sheet_name]
                elif sheet_name == "de_sheet":
                    # Use all keys from the data for DE sheet
                    de_headers = list(data.keys())
                    self.apply_excel_template(sheet_name, de_headers)
                    # Reload the workbook
                    wb = load_workbook(self.full_excel_path)
                    ws = wb[sheet_name]
                else:
                    logger.error(f"Unable to create headers for unknown sheet type: {sheet_name}")
                    return False
            else:
                ws = wb[sheet_name]
            
            # Get the headers from the first row
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(cell.value)
            
            if not headers:
                logger.error(f"No headers found in sheet '{sheet_name}'")
                return False
            
            # Find the next empty row
            next_row = ws.max_row + 1
            
            # Fill in the data
            for col_idx, header in enumerate(headers, start=1):
                # Check if the header exists in the data
                value = data.get(header, "")
                
                # Set the cell value
                cell = ws.cell(row=next_row, column=col_idx, value=value)
                
                # Apply formatting based on content
                if header == "Title":
                    # Make title bold
                    cell.font = Font(bold=True)
                elif "SCORE" in header:
                    # Right-align score values
                    cell.alignment = Alignment(horizontal="right")
                    # Format as number with 1 decimal place if it's a number
                    try:
                        if isinstance(value, (int, float)):
                            cell.number_format = "0.0"
                    except:
                        pass
            
            # Save the workbook with error handling
            try:
                wb.save(self.full_excel_path)
                logger.info(f"Data added to sheet '{sheet_name}' at row {next_row}")
                return True
            except Exception as save_err:
                logger.error(f"Error saving workbook: {save_err}")
                # Try creating a new file with a different name as a last resort
                recovery_path = f"{self.excel_dir}/recovery_{__import__('datetime').datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                try:
                    wb.save(recovery_path)
                    logger.info(f"Saved data to recovery file: {recovery_path}")
                    return True
                except:
                    logger.error("Failed to save recovery file")
                    return False
                
        except Exception as e:
            logger.error(f"Error filling Excel with data for sheet '{sheet_name}': {e}")
            return False
    
    def get_existing_data(self, sheet_name):
        """
        Retrieve existing data from a sheet with better error handling
        """
        try:
            # Ensure we have a valid workbook
            wb = self._ensure_valid_workbook()
            
            # Check if the sheet exists
            if sheet_name not in wb.sheetnames:
                logger.warning(f"Sheet '{sheet_name}' does not exist in the workbook")
                return []
                
            ws = wb[sheet_name]
            
            # Get the headers from the first row
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(cell.value)
            
            if not headers:
                logger.warning(f"No headers found in sheet '{sheet_name}'")
                return []
            
            # Extract the data from each row
            data = []
            for row in range(2, ws.max_row + 1):
                row_data = {}
                for col, header in enumerate(headers, start=1):
                    cell_value = ws.cell(row=row, column=col).value
                    row_data[header] = cell_value
                data.append(row_data)
                
            return data
        except Exception as e:
            logger.error(f"Error retrieving data from sheet '{sheet_name}': {e}")
            return []